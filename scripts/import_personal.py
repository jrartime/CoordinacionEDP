from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\Tbl_Personal.xlsx")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"

HEADER_MAP = {
    "id": "id",
    "id_personal": "id",
    "activo": "activo",
    "pert_empresa": "pert_empresa",
    "vinculacion_id": "vinculacion_id",
    "personal": "personal",
    "genero": "genero",
    "antiguedad": "antiguedad",
    "dni": "dni",
    "fecha_nacimiento": "fecha_nacimiento",
    "ss": "ss",
    "email": "email",
    "movil": "movil",
    "telefono": "telefono",
    "direccion": "direccion",
    "codigo_postal": "codigo_postal",
    "localidad": "localidad",
    "municipio": "municipio",
    "provincia": "provincia",
    "Cuenta_corriente": "cuenta_corriente",
    "foto": "foto",
    "contrato_id": "contrato_id",
    "observacion": "observacion",
    "desplazamiento": "desplazamiento",
    "enviar": "enviar",
    "carpeta": "carpeta",
    "pago": "pago",
    "CV": "cv",
    "DA": "da",
    "DS": "ds",
    "Prev_Riesgos": "prev_riesgos",
    "epi": "epi",
    "Titulos": "titulos",
    "Ig_Ac": "ig_ac",
    "Uniforme": "uniforme",
    "Med_Emerg": "med_emerg",
    "ens": "ens",
    "prorrateo_pagas": "prorrateo_pagas",
    "num_pagas_extra": "num_pagas_extra",
    "tipo_contrato": "tipo_contrato",
    "grupo": "grupo",
    "nivel": "nivel",
    "grupo_cotizacion": "grupo_cotizacion",
    "contacto_urgencia": "contacto_urgencia",
    "telefono_urgencia": "telefono_urgencia",
    "persona": "persona",
    "IRPF": "irpf",
    "nombre": "nombre",
    "apellido": "apellido",
}

INTEGER_FIELDS = {
    "id",
    "vinculacion_id",
    "codigo_postal",
    "num_pagas_extra",
    "tipo_contrato",
    "grupo",
    "nivel",
}
NUMERIC_FIELDS = {"irpf"}
TEXT_FIELDS = {
    "personal",
    "genero",
    "dni",
    "ss",
    "email",
    "movil",
    "telefono",
    "direccion",
    "localidad",
    "municipio",
    "provincia",
    "cuenta_corriente",
    "foto",
    "contrato_id",
    "observacion",
    "carpeta",
    "grupo_cotizacion",
    "contacto_urgencia",
    "telefono_urgencia",
    "nombre",
    "apellido",
}
BOOLEAN_FIELDS = {
    "activo",
    "pert_empresa",
    "desplazamiento",
    "enviar",
    "pago",
    "cv",
    "da",
    "ds",
    "epi",
    "titulos",
    "ig_ac",
    "uniforme",
    "med_emerg",
    "ens",
    "prorrateo_pagas",
    "persona",
}
DATE_FIELDS = {"antiguedad", "fecha_nacimiento", "prev_riesgos"}


def normalize_header(value: object) -> str:
    return str(value or "").strip()


def normalize_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False

    text = str(value).strip().lower()
    if not text:
        return False
    if text in {"true", "t", "1", "si", "sí", "s", "yes", "y", "verdadero"}:
        return True
    if text in {"false", "f", "0", "no", "n", "falso"}:
        return False

    raise ValueError(f"Valor booleano no reconocido: {value!r}")


def normalize_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    raise ValueError(f"Fecha no reconocida: {value!r}")


def split_personal_name(value: object) -> tuple[str | None, str | None]:
    parts = str(value or "").strip().split()
    if not parts:
        return None, None
    if len(parts) == 1:
        return parts[0], None
    if len(parts) == 2:
        return parts[0], parts[1]
    if len(parts) == 3 and parts[1].lower() in {"jesus", "jesús"}:
        return " ".join(parts[:2]), parts[2]
    return " ".join(parts[:-2]), " ".join(parts[-2:])


def normalize_value(column: str, value: object) -> object:
    if column in BOOLEAN_FIELDS:
        return normalize_boolean(value)
    if column in DATE_FIELDS:
        return normalize_date(value)
    if value is None:
        return None
    if column in INTEGER_FIELDS:
        text = str(value).strip()
        return int(Decimal(text.replace(",", "."))) if text else None
    if column in NUMERIC_FIELDS:
        text = str(value).strip().replace(",", ".")
        return str(Decimal(text)) if text else None
    if column in TEXT_FIELDS:
        text = str(value).strip()
        return text or None
    return value


def map_headers(source_headers: list[object]) -> list[str]:
    normalized_headers = [normalize_header(header) for header in source_headers]
    missing = [header for header in normalized_headers if header not in HEADER_MAP]
    if missing:
        raise ValueError(f"Columnas no mapeadas: {missing}")

    return [HEADER_MAP[header] for header in normalized_headers]


def normalize_row(source_row: list[object], target_headers: list[str]) -> dict[str, object]:
    row = {
        column: normalize_value(column, value)
        for column, value in zip(target_headers, source_row)
    }
    if not row.get("id") or not row.get("personal"):
        raise ValueError(f"Fila sin id o personal: {row}")

    if not row.get("nombre") or not row.get("apellido"):
        nombre, apellido = split_personal_name(row.get("personal"))
        row["nombre"] = row.get("nombre") or nombre
        row["apellido"] = row.get("apellido") or apellido

    return row


def load_xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    source_headers = list(next(rows))
    target_headers = map_headers(source_headers)
    loaded_rows: list[dict[str, object]] = []

    for source_row in rows:
        if not any(value is not None for value in source_row):
            continue
        loaded_rows.append(normalize_row(list(source_row), target_headers))

    return loaded_rows


def load_csv_rows(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csvfile:
        sample = csvfile.read(4096)
        csvfile.seek(0)
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(csvfile, delimiter=delimiter)
        source_headers = next(reader)
        target_headers = map_headers(source_headers)
        loaded_rows = []

        for source_row in reader:
            if not any(str(value).strip() for value in source_row):
                continue
            loaded_rows.append(normalize_row(source_row, target_headers))

    return loaded_rows


def load_rows(path: Path) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        return load_csv_rows(path)
    return load_xlsx_rows(path)


def upsert_rows(supabase_url: str, service_role_key: str, rows: list[dict[str, object]]) -> int:
    payload = json.dumps(rows, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        f"{supabase_url.rstrip('/')}/rest/v1/personal?on_conflict=id",
        data=payload,
        method="POST",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=representation",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            body = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8")
        raise RuntimeError(f"Supabase devolvio {exc.code}: {detail}") from exc

    return len(json.loads(body)) if body else 0


def build_sql(rows: list[dict[str, object]], *, fill_missing: bool = False) -> str:
    columns = list(dict.fromkeys(HEADER_MAP.values()))
    type_map = {
        **{column: "integer" for column in INTEGER_FIELDS},
        **{column: "numeric" for column in NUMERIC_FIELDS},
        **{column: "text" for column in TEXT_FIELDS},
        **{column: "boolean" for column in BOOLEAN_FIELDS},
        **{column: "date" for column in DATE_FIELDS},
    }
    select_columns = ", ".join(columns)
    record_columns = ",\n  ".join(f"{column} {type_map[column]}" for column in columns)
    payload = json.dumps(rows, ensure_ascii=False, indent=2)

    if fill_missing:
        fillable_columns = [
            column
            for column in columns
            if column != "id" and column not in BOOLEAN_FIELDS
        ]
        assignments = []
        for column in fillable_columns:
            if column in TEXT_FIELDS:
                assignments.append(
                    f"{column} = case when nullif(btrim(p.{column}), '') is null then s.{column} else p.{column} end"
                )
            else:
                assignments.append(f"{column} = coalesce(p.{column}, s.{column})")

        update_clause = ",\n  ".join(assignments)
        insert_values = ", ".join(f"s.{column}" for column in columns)

        return f"""begin;

create temporary table import_personal_source (
  {record_columns}
) on commit drop;

insert into import_personal_source ({select_columns})
select {select_columns}
from jsonb_to_recordset($personal$
{payload}
$personal$::jsonb) as source (
  {record_columns}
);

create temporary table import_personal_matches on commit drop as
with existing_dni as (
  select upper(btrim(dni)) as dni_key, min(id) as id, count(*) as total
  from public.personal
  where nullif(btrim(dni), '') is not null
  group by upper(btrim(dni))
)
select
  s.*,
  coalesce(p_by_id.id, case when existing_dni.total = 1 then existing_dni.id end) as match_id,
  p_by_id.id is not null as matched_by_id,
  p_by_id.id is null and existing_dni.total = 1 as matched_by_dni,
  p_by_id.id is null and existing_dni.total > 1 as ambiguous_dni
from import_personal_source s
left join public.personal p_by_id
  on p_by_id.id = s.id
left join existing_dni
  on existing_dni.dni_key = upper(btrim(s.dni));

create temporary table import_personal_updated (id integer) on commit drop;

with updated as (
update public.personal p
set
  {update_clause}
from import_personal_matches s
where p.id = s.match_id
returning p.id
)
insert into import_personal_updated (id)
select id
from updated;

create temporary table import_personal_inserted (id integer) on commit drop;

with inserted as (
insert into public.personal ({select_columns})
select {insert_values}
from import_personal_matches s
where s.match_id is null
  and not s.ambiguous_dni
on conflict (id) do nothing
returning id
)
insert into import_personal_inserted (id)
select id
from inserted;

select
  (select count(*) from import_personal_source) as filas_origen,
  (select count(*) from import_personal_matches where matched_by_id) as existentes_por_id,
  (select count(*) from import_personal_matches where matched_by_dni) as existentes_por_dni,
  (select count(*) from import_personal_updated) as filas_revisadas_para_rellenar,
  (select count(*) from import_personal_inserted) as filas_insertadas,
  (select count(*) from import_personal_matches where ambiguous_dni) as dni_ambiguos_no_insertados;

select id, personal, dni
from import_personal_matches
where ambiguous_dni
order by personal;

commit;
"""

    update_clause = ",\n  ".join(
        f"{column} = excluded.{column}" for column in columns if column != "id"
    )

    return f"""insert into public.personal ({select_columns})
select {select_columns}
from jsonb_to_recordset($personal$
{payload}
$personal$::jsonb) as source (
  {record_columns}
)
on conflict (id) do update set
  {update_clause};
"""


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Importa Tbl_Personal.xlsx en public.personal.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    parser.add_argument("--output", type=Path)
    parser.add_argument("--print-sql", action="store_true")
    parser.add_argument(
        "--fill-missing",
        action="store_true",
        help="Genera SQL conservador: inserta nuevos y solo rellena campos nulos/vacios en existentes.",
    )
    args = parser.parse_args()

    rows = load_rows(args.source)

    if args.print_sql or args.output:
        sql = build_sql(rows, fill_missing=args.fill_missing)
        if args.output:
            args.output.parent.mkdir(parents=True, exist_ok=True)
            args.output.write_text(sql, encoding="utf-8")
            print(f"SQL generado en {args.output}")
        else:
            print(sql)
        return

    if args.fill_missing:
        raise SystemExit("--fill-missing solo esta disponible con --print-sql o --output.")

    if not args.service_role_key:
        raise SystemExit("Falta SUPABASE_SERVICE_ROLE_KEY o --service-role-key para escribir en Supabase.")

    count = upsert_rows(args.url, args.service_role_key, rows)
    print(f"Registros insertados/actualizados: {count}")


if __name__ == "__main__":
    main()

"""Importa el historico tbl_horas (Excel, formato tbl_horas2021.xlsx) a public.registros.

Reglas aplicadas (acordadas para la migracion del modelo de horas por apuntes):
  - horas_final = primer valor no nulo entre horas, hc, hf, hm (convencion "ancha"
    legacy: para MONT/HCOMP/FTRAB la magnitud vive en su columna propia, no en horas).
  - Si horas_final sigue sin dato pero hora_inicio y hora_fin estan presentes,
    se reconstruye como fin - inicio (cruzando medianoche si hace falta).
  - Si nada de lo anterior aplica, horas_final = 0.
  - situacion_id = 9 (codigo inexistente en el catalogo actual) -> se importa como NULL.
  - Se descartan (no se cargan) las filas con horas negativas o con FTRAB donde
    horas y hf estan ambos presentes y no coinciden (ambiguas). Van a un CSV de
    descartes para revision manual.
  - clases, hd, bolsa_horas, horas_2, titular_personal_id, sustituto_personal_id
    e institucion_id del Excel NO se migran (campos legacy fuera del modelo actual).
  - actividad_id del Excel NO se migra (se importa NULL): public.actividades es la
    tabla de planificacion ACTUAL (822 filas, ids 1-922) y ninguno de los 2.808
    actividad_id distintos del historico (rango 0-11610, anios 2021-2026) coincide
    con una actividad viva -> el 100% violaria la FK si se intentara preservar.
  - id_hora se conserva en registros.legacy_id_hora (columna de auditoria, no
    usada por la app) para trazabilidad; el insert usa upsert por legacy_id_hora
    para que volver a ejecutar el script no duplique filas.

Uso:
  python scripts/import_registros_historico.py --dry-run
  python scripts/import_registros_historico.py --dry-run --source ruta/a/muestra.xlsx
  python scripts/import_registros_historico.py --service-role-key ...
  (o con SUPABASE_SERVICE_ROLE_KEY en el entorno)
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, time as dtime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\tbl_horas2021.xlsx")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"
DEFAULT_BATCH_SIZE = 5000

EXPECTED_HEADERS = [
    "id_hora", "fecha", "actividad_id", "empresa_id", "contrato_id", "personal_id",
    "titular_personal_id", "sustituto_personal_id", "instalacion_id", "institucion_id",
    "categoria_id", "puesto_id", "funcion_id", "modalidad_id", "Nivel_id", "grupo_id",
    "Nota", "dia_id", "hora_inicio", "hora_fin", "horas", "hc", "hf", "hm", "hd",
    "bolsa_horas", "horas_diurnas", "horas_nocturnas", "clases", "horas_2", "descanso",
    "activo", "festivo", "sustitucion", "facturar", "abonar", "tipo_hora_id",
    "situacion_id", "Año", "observacion", "control", "factura",
]

# Codigos de situacion que ya no existen en el catalogo public.situaciones y que
# se importan como NULL en vez de romper la fila.
SITUACION_UNKNOWN = {9}

TARGET_COLUMNS = [
    "legacy_id_hora", "fecha", "actividad_id", "empresa_id", "contrato_id", "personal_id",
    "instalacion_id", "categoria_id", "puesto_id", "funcion_id", "modalidad_id",
    "nivel_id", "grupo_id", "nota", "dia_id", "hora_inicio", "hora_fin", "horas",
    "horas_diurnas", "horas_nocturnas", "descanso", "activo", "festivo", "sustitucion",
    "facturar", "abonar", "tipo_hora_id", "situacion_id", "anio", "observacion",
    "control", "factura",
]


def to_iso_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def to_iso_time(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, dtime):
        return value.strftime("%H:%M:%S")
    return str(value)


def to_iso_datetime(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def to_bool(value: object) -> bool:
    return bool(value) if value is not None else False


def to_num(value: object) -> float | None:
    if value is None:
        return None
    return float(value)


def to_int(value: object) -> int | None:
    if value is None:
        return None
    return int(value)


def compute_horas_final(horas, hc, hf, hm, hora_inicio, hora_fin) -> tuple[float, str]:
    """Devuelve (valor, origen) segun la regla de coalescencia acordada."""
    for source_name, candidate in (("horas", horas), ("hc", hc), ("hf", hf), ("hm", hm)):
        if candidate is not None:
            return float(candidate), source_name

    if isinstance(hora_inicio, dtime) and isinstance(hora_fin, dtime):
        start = hora_inicio.hour * 3600 + hora_inicio.minute * 60 + hora_inicio.second
        end = hora_fin.hour * 3600 + hora_fin.minute * 60 + hora_fin.second
        if end < start:
            end += 86400
        return round((end - start) / 3600.0, 2), "reconstruido_horario"

    return 0.0, "sin_dato"


def load_and_transform(path: Path, limit: int | None = None):
    """Devuelve (filas_validas, filas_descartadas) ya normalizadas."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else h for h in next(rows)]

    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        raise ValueError(f"Faltan columnas esperadas en el Excel: {missing}")

    idx = {h: i for i, h in enumerate(headers)}
    valid_rows: list[dict[str, object]] = []
    discarded_rows: list[dict[str, object]] = []

    for n, row in enumerate(rows):
        if row[idx["id_hora"]] is None:
            continue
        if limit is not None and n >= limit:
            break

        id_hora = row[idx["id_hora"]]
        horas_raw = row[idx["horas"]]
        hc_raw = row[idx["hc"]]
        hf_raw = row[idx["hf"]]
        hm_raw = row[idx["hm"]]
        tipo_hora_id = to_int(row[idx["tipo_hora_id"]])
        hora_inicio = row[idx["hora_inicio"]]
        hora_fin = row[idx["hora_fin"]]

        # Caso B: horas negativas -> descartar (dato anomalo, no auto-resoluble).
        if horas_raw is not None and float(horas_raw) < 0:
            discarded_rows.append({
                "id_hora": id_hora, "motivo": "horas_negativas",
                "detalle": f"horas={horas_raw}",
            })
            continue

        # Caso D: FTRAB con horas y hf presentes y distintos -> ambiguo, descartar.
        if tipo_hora_id == 4 and horas_raw is not None and hf_raw is not None:
            if abs(float(horas_raw) - float(hf_raw)) > 0.01:
                discarded_rows.append({
                    "id_hora": id_hora, "motivo": "ftrab_horas_hf_ambiguo",
                    "detalle": f"horas={horas_raw} hf={hf_raw}",
                })
                continue

        horas_final, origen = compute_horas_final(
            to_num(horas_raw), to_num(hc_raw), to_num(hf_raw), to_num(hm_raw),
            hora_inicio, hora_fin,
        )

        situacion_id = to_int(row[idx["situacion_id"]])
        if situacion_id in SITUACION_UNKNOWN:
            situacion_id = None

        valid_rows.append({
            "legacy_id_hora": to_int(id_hora),
            "fecha": to_iso_date(row[idx["fecha"]]),
            # actividad_id no se migra: la tabla actividades es la planificacion
            # actual, ningun id historico coincide (ver docstring del modulo).
            "actividad_id": None,
            "empresa_id": to_int(row[idx["empresa_id"]]) or 1,
            "contrato_id": to_int(row[idx["contrato_id"]]),
            "personal_id": to_int(row[idx["personal_id"]]),
            "instalacion_id": to_int(row[idx["instalacion_id"]]),
            "categoria_id": to_int(row[idx["categoria_id"]]),
            "puesto_id": to_int(row[idx["puesto_id"]]),
            "funcion_id": to_int(row[idx["funcion_id"]]),
            "modalidad_id": to_int(row[idx["modalidad_id"]]),
            "nivel_id": to_int(row[idx["Nivel_id"]]),
            "grupo_id": to_int(row[idx["grupo_id"]]),
            "nota": row[idx["Nota"]],
            "dia_id": to_int(row[idx["dia_id"]]),
            "hora_inicio": to_iso_time(hora_inicio),
            "hora_fin": to_iso_time(hora_fin),
            "horas": horas_final,
            "horas_diurnas": to_num(row[idx["horas_diurnas"]]),
            "horas_nocturnas": to_num(row[idx["horas_nocturnas"]]),
            "descanso": to_bool(row[idx["descanso"]]),
            "activo": to_bool(row[idx["activo"]]),
            "festivo": to_bool(row[idx["festivo"]]),
            "sustitucion": to_bool(row[idx["sustitucion"]]),
            "facturar": to_bool(row[idx["facturar"]]) if row[idx["facturar"]] is not None else True,
            "abonar": to_bool(row[idx["abonar"]]) if row[idx["abonar"]] is not None else True,
            "tipo_hora_id": tipo_hora_id,
            "situacion_id": situacion_id,
            "anio": to_int(row[idx["Año"]]),
            "observacion": row[idx["observacion"]],
            "control": to_iso_datetime(row[idx["control"]]),
            "factura": row[idx["factura"]],
            "_horas_origen": origen,
        })

    return valid_rows, discarded_rows


def write_discard_report(path: Path, discarded_rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["id_hora", "motivo", "detalle"])
        writer.writeheader()
        for row in discarded_rows:
            writer.writerow(row)


def post_batch(supabase_url: str, service_role_key: str, batch: list[dict[str, object]]) -> None:
    payload = json.dumps(
        [{k: v for k, v in row.items() if not k.startswith("_")} for row in batch],
        ensure_ascii=False,
    ).encode("utf-8")
    request = urllib.request.Request(
        f"{supabase_url.rstrip('/')}/rest/v1/registros?on_conflict=legacy_id_hora",
        data=payload,
        method="POST",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            response.read()
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8")
        raise RuntimeError(f"Supabase devolvio {exc.code}: {detail}") from exc


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Importa el historico tbl_horas a public.registros.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--limit", type=int, default=None, help="Procesa solo las primeras N filas (pruebas).")
    parser.add_argument("--discard-report", type=Path, default=Path("exports/registros_historico_descartes.csv"))
    parser.add_argument("--dry-run", action="store_true", help="Solo analiza y reporta, no escribe en Supabase.")
    args = parser.parse_args()

    print(f"Leyendo {args.source} ...")
    valid_rows, discarded_rows = load_and_transform(args.source, limit=args.limit)

    origen_counts: dict[str, int] = {}
    for row in valid_rows:
        origen_counts[row["_horas_origen"]] = origen_counts.get(row["_horas_origen"], 0) + 1

    print(f"Filas totales procesadas: {len(valid_rows) + len(discarded_rows)}")
    print(f"  Validas:    {len(valid_rows)}")
    print(f"  Descartadas: {len(discarded_rows)}")
    print("Origen del valor de horas en las filas validas:")
    for origen, count in sorted(origen_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {origen}: {count}")

    if discarded_rows:
        write_discard_report(args.discard_report, discarded_rows)
        print(f"Informe de descartes escrito en {args.discard_report}")

    if args.dry_run:
        print("Dry-run: no se escribio nada en Supabase.")
        return

    if not args.service_role_key:
        raise SystemExit("Falta SUPABASE_SERVICE_ROLE_KEY o --service-role-key para escribir en Supabase.")

    total = len(valid_rows)
    for start in range(0, total, args.batch_size):
        batch = valid_rows[start:start + args.batch_size]
        post_batch(args.url, args.service_role_key, batch)
        print(f"  Cargadas {min(start + len(batch), total)}/{total}")
        time.sleep(0.2)

    print("Carga completada.")


if __name__ == "__main__":
    main()

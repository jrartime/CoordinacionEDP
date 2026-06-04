from __future__ import annotations

import csv
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook


SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\Documents\Tbl_Registros.xlsx")
OUTPUT_PATH = Path(r"D:\Programación\CurriculosEDP\exports\tbl_registros.csv")

HEADER_MAP = {
    "id": "id",
    "Personal": "personal",
    "DNI": "dni",
    "Centro": "centro",
    "Puesto": "puesto",
    "Fecha": "fecha",
    "Hora_inicio": "hora_inicio",
    "Hora_fin": "hora_fin",
    "Tipo_jornada": "tipo_jornada",
    "Observacion": "observacion",
    "Eliminado": "eliminado",
    "contriol": "control",
}


def normalize_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    source_headers = next(rows)
    target_headers = [HEADER_MAP[h] for h in source_headers]

    with OUTPUT_PATH.open("w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(target_headers)
        for row in rows:
            normalized = []
            for mapped, value in zip(target_headers, row):
                if mapped == "fecha" and isinstance(value, datetime):
                    normalized.append(value.date().isoformat())
                else:
                    normalized.append(normalize_value(value))
            writer.writerow(normalized)

    print(f"CSV generado en {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

import argparse
import csv
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook


TARGET_COLUMNS = [
    "personal_id",
    "contrato_id",
    "servicio_id",
    "empresa_id",
    "instalacion_id",
    "puesto_id",
    "funcion_id",
    "modalidad_id",
    "situacion_id",
    "tipo_hora_id",
    "dias_semana",
    "fecha_inicio",
    "fecha_fin",
    "hora_inicio",
    "hora_fin",
    "llamamiento_enviado",
    "respuesta_llamamiento",
    "observaciones",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Convierte el CSV antiguo de actividades de conserjes al esquema de Supabase."
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("days", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--times",
        type=Path,
        required=True,
        help="XLSX original que conserva hora_inicio y hora_fin correctamente.",
    )
    parser.add_argument(
        "--services",
        type=Path,
        help="CSV convertido anterior del que recuperar servicio_id por actividad equivalente.",
    )
    return parser.parse_args()


def parse_date(value):
    return datetime.strptime(value.strip(), "%d/%m/%Y").date().isoformat()


def format_time(value):
    if not isinstance(value, time):
        raise ValueError(f"Hora no valida en el XLSX: {value!r}")
    return value.strftime("%H:%M:%S")


def load_days(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    id_index = headers.index("id_dia")
    number_index = headers.index("dia_numero")
    result = {}
    for row in rows:
        if row[id_index] is None:
            continue
        result[str(row[id_index]).strip()] = [
            int(character) for character in str(row[number_index]).strip()
        ]
    return result


def activity_key(row):
    return (
        str(row["personal_id"]).strip(),
        str(row["contrato_id"]).strip(),
        str(row["empresa_id"]).strip(),
        str(row["instalacion_id"]).strip(),
        str(row["puesto_id"]).strip(),
        str(row["situacion_id"]).strip(),
        str(row["tipo_hora_id"]).strip(),
        str(row["fecha_inicio"]).strip(),
        str(row["fecha_fin"]).strip(),
        str(row["dias_semana"]).strip(),
    )


def load_services(path):
    if not path:
        return {}
    with path.open(encoding="utf-8-sig", newline="") as source:
        rows = list(csv.DictReader(source))
    return {
        activity_key(row): str(row.get("servicio_id") or "").strip()
        for row in rows
    }


def load_times(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    result = {}
    for row in rows:
        source = dict(zip(headers, row))
        result[str(source["id"]).strip()] = (
            format_time(source["hora_inicio"]),
            format_time(source["hora_fin"]),
        )
    return result


def convert_rows(source_path, day_map, service_map, time_map):
    converted = []
    discarded = []
    with source_path.open(encoding="utf-8-sig", newline="") as source:
        rows = csv.DictReader(source, delimiter=";")
        for source_row in rows:
            if source_row["activo"].strip().upper() != "VERDADERO":
                discarded.append(source_row["id"].strip())
                continue

            days = day_map[source_row["dia_id"].strip()]
            start_time, end_time = time_map[source_row["id"].strip()]
            observations = "; ".join(
                value.strip()
                for value in (source_row.get("observacion", ""), source_row.get("sala", ""))
                if value and value.strip()
            )
            target = {
                "personal_id": source_row["personal_id"].strip(),
                "contrato_id": source_row["contrato_id"].strip(),
                "servicio_id": "",
                "empresa_id": source_row["empresa_id"].strip(),
                "instalacion_id": source_row["instalacion_id"].strip(),
                "puesto_id": source_row["puesto_id"].strip(),
                "funcion_id": source_row["funcion_id"].strip(),
                "modalidad_id": source_row["modalidad_id"].strip(),
                "situacion_id": source_row["situacion_id"].strip(),
                "tipo_hora_id": source_row["tipo_hora_id"].strip(),
                "dias_semana": "{" + ",".join(map(str, days)) + "}",
                "fecha_inicio": parse_date(source_row["fecha_alta"]),
                "fecha_fin": parse_date(source_row["fecha_baja"]),
                "hora_inicio": start_time,
                "hora_fin": end_time,
                "llamamiento_enviado": "false",
                "respuesta_llamamiento": "",
                "observaciones": observations,
            }
            target["servicio_id"] = service_map.get(activity_key(target), "")
            converted.append(target)
    return converted, discarded


def main():
    args = parse_args()
    day_map = load_days(args.days)
    service_map = load_services(args.services)
    time_map = load_times(args.times)
    converted, discarded = convert_rows(args.source, day_map, service_map, time_map)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=TARGET_COLUMNS)
        writer.writeheader()
        writer.writerows(converted)
    print(f"Generadas {len(converted)} actividades activas en {args.output}")
    print(f"Descartadas {len(discarded)} actividades inactivas: {', '.join(discarded)}")


if __name__ == "__main__":
    main()

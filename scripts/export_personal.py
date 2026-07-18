from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\Documents\Tablas maestras\Tbl_Personal.xlsx")
OUTPUT_PATH = Path(r"D:\Programación\CurriculosEDP\exports\personal.csv")

HEADER_MAP = {
    "id_personal": "id",
    "activo": "activo",
    "pert_empresa": "pert_empresa",
    "vinculacion_id": "vinculacion_id",
    "personal": "personal",
    "genero": "genero",
    "antiguedad": "antiguedad",
    "dni": "dni",
    "fecha_nacimiento": "fecha_nacimiento",
    "ss": "ss",
    "email": "email",
    "movil": "movil",
    "telefono": "telefono",
    "direccion": "direccion",
    "codigo_postal": "codigo_postal",
    "localidad": "localidad",
    "municipio": "municipio",
    "provincia": "provincia",
    "Cuenta_corriente": "cuenta_corriente",
    "foto": "foto",
    "contrato_id": "contrato_id",
    "observacion": "observacion",
    "desplazamiento": "desplazamiento",
    "enviar": "enviar",
    "carpeta": "carpeta",
    "pago": "pago",
    "CV": "cv",
    "DA": "da",
    "DS": "ds",
    "Prev_Riesgos": "prev_riesgos",
    "epi": "epi",
    "Titulos": "titulos",
    "Ig_Ac": "ig_ac",
    "Uniforme": "uniforme",
    "Med_Emerg": "med_emerg",
    "ens": "ens",
    "prorrateo_pagas": "prorrateo_pagas",
    "num_pagas_extra": "num_pagas_extra",
    "tipo_contrato": "tipo_contrato",
    "grupo": "grupo",
    "nivel": "nivel",
    "grupo_cotizacion": "grupo_cotizacion",
    "contacto_urgencia": "contacto_urgencia",
    "telefono_urgencia": "telefono_urgencia",
    "persona": "persona",
    "IRPF": "irpf",
    "nombre": "nombre",
    "apellido": "apellido",
}

NUMERIC_2_FIELDS = set()
NUMERIC_4_FIELDS = {"irpf"}


def normalize_value(column: str, value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if column in NUMERIC_2_FIELDS:
        return format(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")
    if column in NUMERIC_4_FIELDS:
        return format(Decimal(str(value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP), "f")
    return value


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    source_headers = next(rows)
    target_headers = [HEADER_MAP[h] for h in source_headers]

    with OUTPUT_PATH.open("w", newline="", encoding="utf-8-sig") as csvfile:
      writer = csv.writer(csvfile)
      writer.writerow(target_headers)
      for row in rows:
          writer.writerow(
              [normalize_value(mapped, value) for mapped, value in zip(target_headers, row)]
          )

    print(f"CSV generado en {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

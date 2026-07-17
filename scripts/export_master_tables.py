from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


BASE_SOURCE = Path(r"C:\Users\Jr\OneDrive\Documents\Tablas maestras")
BASE_OUTPUT = Path(r"D:\Programación\CurriculosEDP\exports")

TABLES = {
    "contratos": {
        "source": "contratos.xlsx",
        "header_map": {
            "id_contrato": "id",
            "contrato": "contrato",
            "descripcion": "descripcion",
            "presupuesto_anual": "presupuesto_anual",
            "fecha_inicio": "fecha_inicio",
            "fecha_fin": "fecha_fin",
            "expediente": "expediente",
            "CPV": "cpv",
            "importe": "importe",
            "cliente": "cliente",
            "activo": "activo",
            "seleccionar": "seleccionar",
            "desplazamiento": "desplazamiento",
            "Agrupacion_nomina": "agrupacion_nomina",
            "IVA": "iva",
        },
        "numeric_2": {"presupuesto_anual"},
        "numeric_4": {"iva"},
    },
    "empresas": {
        "source": "empresas.xlsx",
        "header_map": {
            "Id_empresa": "id",
            "empresa": "empresa",
            "razon_social": "razon_social",
            "cif": "cif",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "funciones": {
        "source": "funciones.xlsx",
        "header_map": {
            "id_funcion": "id",
            "funcion": "funcion",
            "siglas": "siglas",
            "grupo": "grupo",
            "formacion_horario": "formacion_horario",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "instalaciones": {
        "source": "instalaciones.xlsx",
        "header_map": {
            "id_instalacion": "id",
            "instalacion": "instalacion",
            "direccion": "direccion",
            "codigo_postal": "codigo_postal",
            "localidad": "localidad",
            "Provincia": "provincia",
            "telefono": "telefono",
            "gps_latitud": "gps_latitud",
            "gps_longitud": "gps_longitud",
            "encargado": None,
            "siglas": "siglas",
            "orden": None,
            "categoria": "categoria",
            "contrato": None,
            "activo": "activo",
            "Activo": "activo",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "modalidades": {
        "source": "modalidades.xlsx",
        "header_map": {
            "id_modalidad": "id",
            "modalidad": "modalidad",
            "siglas": "siglas",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "puestos": {
        "source": "puestos.xlsx",
        "header_map": {
            "id_puesto": "id",
            "puesto": "puesto",
            "detalle_puesto": "detalle_puesto",
            "siglas": "siglas",
            "convenio_grupo_nivel": "convenio_id",
            "categoria_id": "categoria_id",
            "DEA": "dea",
            "Rec_medico": "rec_medico",
            "EPI": "epi",
            "Clausula_preferencia": "clausula_preferencia",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "situaciones": {
        "source": "situaciones.xlsx",
        "header_map": {
            "id_situacion": "id",
            "situacion": "situacion",
            "descripcion": "descripcion",
            "desplazamiento": "desplazamiento",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
    "tipo_horas": {
        "source": "tipo_horas.xlsx",
        "header_map": {
            "id_tipo_hora": "id",
            "tipo_hora": "tipo_hora",
            "descripcion": "descripcion",
        },
        "numeric_2": set(),
        "numeric_4": set(),
    },
}


def normalize_value(column: str, value: object, numeric_2: set[str], numeric_4: set[str]) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if column in numeric_2:
        return format(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")
    if column in numeric_4:
        return format(Decimal(str(value)).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP), "f")
    return value


def export_table(table_name: str, config: dict) -> None:
    source_path = BASE_SOURCE / config["source"]
    output_path = BASE_OUTPUT / f"{table_name}.csv"

    wb = load_workbook(source_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    source_headers = next(rows)
    mapped_columns = [
        (source_index, config["header_map"][header])
        for source_index, header in enumerate(source_headers)
        if config["header_map"][header] is not None
    ]
    target_headers = [target_header for _, target_header in mapped_columns]

    with output_path.open("w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(target_headers)
        for row in rows:
            writer.writerow(
                [
                    normalize_value(mapped, row[source_index], config["numeric_2"], config["numeric_4"])
                    for source_index, mapped in mapped_columns
                ]
            )

    print(f"CSV generado en {output_path}")


def main() -> None:
    BASE_OUTPUT.mkdir(parents=True, exist_ok=True)
    for table_name, config in TABLES.items():
        export_table(table_name, config)


if __name__ == "__main__":
    main()

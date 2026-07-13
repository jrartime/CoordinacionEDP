"""Importa CronosBanco.xlsx (movimientos bancarios/TPV) a public.cronos_banco.

Origen: hoja "Banco", 12 columnas en orden fijo (se mapean por POSICION para no
depender de la codificacion exacta de las cabeceras):
  Id, Fecha, Hora, Nº terminal, Tipo operacion, Cod pedido,
  Resultado operacion y codigo, Importe, Moneda, Importe Euros,
  Tipo de pago, Nº tarjeta

  - Fecha (datetime) -> date ISO; Hora (time) -> HH:MM:SS.
  - Importe / Importe Euros -> numeric (admite coma decimal en texto).
  - "Id" unico -> id_origen (upsert idempotente por id_origen).

Uso:
  python scripts/import_cronos_banco.py --dry-run
  python scripts/import_cronos_banco.py --service-role-key ...
  (o con SUPABASE_SERVICE_ROLE_KEY en el entorno)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, time as dtime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\CronosBanco.xlsx")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"
DEFAULT_BATCH_SIZE = 5000
EMPTY_TOKENS = {"", "----"}


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return None if text in EMPTY_TOKENS else text


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = clean(value)
    if text and "/" in text:
        day, month, year = text.split("/")
        return f"{year.zfill(4)}-{month.zfill(2)}-{day.zfill(2)}"
    return None  # "--" u otros marcadores no-fecha -> NULL


def to_iso_time(value):
    if value is None:
        return None
    if isinstance(value, (dtime, datetime)):
        return value.strftime("%H:%M:%S")
    return clean(value)


def to_num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if text is None:
        return None
    return float(text.replace(".", "").replace(",", ".")) if "," in text else float(text)


def to_int(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = clean(value)
    if text is None:
        return None
    return int(float(text.replace(",", ".")))


# Columnas de destino en el ORDEN de las columnas del Excel (mapeo posicional).
COLUMNS = [
    ("id_origen", to_int),
    ("fecha", to_iso_date),
    ("hora", to_iso_time),
    ("terminal", to_int),
    ("tipo_operacion", clean),
    ("cod_pedido", to_int),
    ("resultado", clean),
    ("importe", to_num),
    ("moneda", clean),
    ("importe_euros", to_num),
    ("tipo_pago", clean),
    ("tarjeta", clean),
]


def load_and_transform(path: Path, limit: int | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    if len(headers) < len(COLUMNS):
        raise ValueError(
            f"El Excel tiene {len(headers)} columnas; se esperaban al menos {len(COLUMNS)}."
        )

    records: list[dict[str, object]] = []
    for n, row in enumerate(rows):
        if limit is not None and n >= limit:
            break
        if row[0] is None:  # sin Id -> fila vacia
            continue
        record = {column: transform(row[index]) for index, (column, transform) in enumerate(COLUMNS)}
        records.append(record)
    return records


def post_batch(supabase_url: str, service_role_key: str, batch: list[dict[str, object]]) -> None:
    payload = json.dumps(batch, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        f"{supabase_url.rstrip('/')}/rest/v1/cronos_banco?on_conflict=id_origen",
        data=payload,
        method="POST",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            response.read()
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8")
        raise RuntimeError(f"Supabase devolvio {exc.code}: {detail}") from exc


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Importa CronosBanco.xlsx a public.cronos_banco.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--limit", type=int, default=None, help="Procesa solo las primeras N filas (pruebas).")
    parser.add_argument("--dry-run", action="store_true", help="Solo analiza, no escribe en Supabase.")
    args = parser.parse_args()

    print(f"Leyendo {args.source} ...")
    rows = load_and_transform(args.source, limit=args.limit)
    print(f"Filas validas: {len(rows)}")
    if rows:
        muestra = {k: rows[0][k] for k in ("id_origen", "fecha", "hora", "importe_euros", "tipo_operacion", "cod_pedido")}
        print(f"  Ejemplo (fila 1): {muestra}")

    if args.dry_run:
        print("Dry-run: no se escribio nada en Supabase.")
        return

    if not args.service_role_key:
        raise SystemExit("Falta SUPABASE_SERVICE_ROLE_KEY o --service-role-key para escribir en Supabase.")

    total = len(rows)
    for start in range(0, total, args.batch_size):
        batch = rows[start:start + args.batch_size]
        post_batch(args.url, args.service_role_key, batch)
        print(f"  Cargadas {min(start + len(batch), total)}/{total}")
        time.sleep(0.2)

    print("Carga completada.")


if __name__ == "__main__":
    main()

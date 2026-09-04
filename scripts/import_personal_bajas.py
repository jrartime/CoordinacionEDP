"""Importa Tbl_Personal_Bajas.xlsx (Access) a public.personal_bajas.

Estrategia: UPSERT por `id`, nunca borrado. Mismo motivo que import_vida_laboral.py:
si se reedita el Excel a mano y se vuelve a correr el script, no debe perder nada
que ya se haya cargado en Supabase.

Filas del Excel excluidas por no cumplir las NOT NULL/CHECK de la tabla (se listan
al ejecutar, no se importan silenciosamente):
  - id 92:  Tipo vacio.
  - id 148 y 150: fecha_inicio vacia.

`baja` (Excel) -> `con_parte_baja`: mismo significado, mapeo directo.
`lugar`/`Tipo` (Excel) -> `lugar_id`/`tipo_id`: los ids del catalogo Access se
preservaron tal cual en personal_bajas_tipo/personal_bajas_lugar (ver
supabase/tables/personal_bajas.sql), no hace falta remapeo.
`ingreso_hospitalario` no existe en el Excel origen: se importa siempre a false;
revisar a mano las filas con "hospitalizado"/"Ingreso Hospitalario" en observacion
si se quiere marcarlas.
`periodo` del Excel no se importa: no cuadraba con fecha_fin - fecha_inicio en unas
25 filas (recaidas o tecleo). dias/en_curso se calculan en personal_bajas_detalle.

Ademas de las filas propias, el script reclasifica 7 filas de
Tbl_Personal_Conciliacion.xlsx (tipo "Medida especial de conciliacion", con
observacion literal "Maternidad"/"Paternidad"/"Riesgo en el Embarazo") hacia esta
tabla, porque el nuevo modelo les da su propio tipo (con sin_nomina_empresa=true) en
vez de forzarlas en conciliacion como hacia Access por no tener mejor sitio. Ver
RECLASIFICAR_DESDE_CONCILIACION. Para que el upsert sea idempotente sin id propio en
origen, se les asigna id = 10000 + id_conciliacion (fuera del rango de
Tbl_Personal_Bajas, que no pasa de 173).

Uso:
  python scripts/import_personal_bajas.py --dry-run
  python scripts/import_personal_bajas.py --service-role-key ...
  (o con SUPABASE_SERVICE_ROLE_KEY en el entorno)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

DEFAULT_BAJAS_SOURCE = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\Tbl_Personal_Bajas.xlsx")
DEFAULT_CONCILIACION_SOURCE = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\Tbl_Personal_Conciliacion.xlsx")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"
DEFAULT_BATCH_SIZE = 500

# Id_personal_bajas -> motivo de exclusion (no cumplen NOT NULL/CHECK).
EXCLUIR = {
    92: "Tipo vacio en el Excel",
    148: "fecha_inicio vacia en el Excel",
    150: "fecha_inicio vacia en el Excel",
}

# Id_personal_conciliacion -> tipo_id de personal_bajas_tipo (6 = Maternidad/paternidad,
# 7 = Riesgo durante el embarazo). Ver import_personal_conciliacion.py, que excluye
# estas mismas filas de su import.
RECLASIFICAR_DESDE_CONCILIACION = {
    45: 6, 46: 6, 57: 6, 64: 6, 66: 7, 68: 6, 71: 6,
}
RECLASIFICAR_ID_OFFSET = 10000


def to_date(v) -> str | None:
    return None if v is None else v.date().isoformat()


def build_records_propias(path: Path) -> tuple[list[dict], list[tuple[int, str]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]

    records = []
    excluidas = []
    for row in rows:
        rid, personal_id, fecha_inicio, fecha_fin, tipo, baja, lugar, _periodo, observacion = row
        if rid in EXCLUIR:
            excluidas.append((rid, EXCLUIR[rid]))
            continue
        records.append({
            "id": rid,
            "personal_id": personal_id,
            "tipo_id": tipo,
            "lugar_id": lugar,
            "fecha_inicio": to_date(fecha_inicio),
            "fecha_fin": to_date(fecha_fin),
            "con_parte_baja": bool(baja),
            "ingreso_hospitalario": False,
            "observacion": observacion,
        })
    return records, excluidas


def build_records_reclasificadas(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]

    records = []
    for row in rows:
        rid, personal_id, fecha_inicio, fecha_fin, tipo, observacion = row
        if rid not in RECLASIFICAR_DESDE_CONCILIACION:
            continue
        nota = f"{observacion} " if observacion else ""
        records.append({
            "id": RECLASIFICAR_ID_OFFSET + rid,
            "personal_id": personal_id,
            "tipo_id": RECLASIFICAR_DESDE_CONCILIACION[rid],
            "lugar_id": None,
            "fecha_inicio": to_date(fecha_inicio),
            "fecha_fin": to_date(fecha_fin),
            "con_parte_baja": True,
            "ingreso_hospitalario": False,
            "observacion": f"{nota}(reclasificado desde personal_conciliacion id {rid})",
        })
    return records


def post_batch(supabase_url: str, service_role_key: str, batch: list[dict]) -> None:
    payload = json.dumps(batch, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        f"{supabase_url.rstrip('/')}/rest/v1/personal_bajas?on_conflict=id",
        data=payload,
        method="POST",
        headers={
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            response.read()
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8")
        raise RuntimeError(f"Supabase devolvio {exc.code}: {detail}") from exc


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="Importa Tbl_Personal_Bajas.xlsx a public.personal_bajas (upsert por id).")
    parser.add_argument("--source", type=Path, default=DEFAULT_BAJAS_SOURCE)
    parser.add_argument("--conciliacion-source", type=Path, default=DEFAULT_CONCILIACION_SOURCE)
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--skip-reclasificacion", action="store_true",
                         help="No incluye las 7 filas de maternidad/paternidad reclasificadas desde conciliacion.")
    parser.add_argument("--dry-run", action="store_true", help="Solo analiza, no escribe en Supabase.")
    args = parser.parse_args()

    print(f"Leyendo {args.source} ...")
    records, excluidas = build_records_propias(args.source)
    print(f"  filas propias: {len(records)}")
    for rid, motivo in excluidas:
        print(f"  EXCLUIDA id {rid}: {motivo}")

    if not args.skip_reclasificacion:
        print(f"Leyendo {args.conciliacion_source} para la reclasificacion ...")
        reclasificadas = build_records_reclasificadas(args.conciliacion_source)
        print(f"  filas reclasificadas desde conciliacion: {len(reclasificadas)}")
        records += reclasificadas

    if args.dry_run:
        print("\n--dry-run: no se escribe nada. Muestra de la primera fila:")
        print(json.dumps(records[0], ensure_ascii=False, indent=2))
        return

    if not args.service_role_key:
        raise SystemExit("Falta --service-role-key (o SUPABASE_SERVICE_ROLE_KEY en el entorno).")

    for i in range(0, len(records), args.batch_size):
        batch = records[i:i + args.batch_size]
        post_batch(args.url, args.service_role_key, batch)
        print(f"  {i + len(batch)}/{len(records)}")

    print(f"\nListo: {len(records)} filas upserted.")
    print(
        "\nPENDIENTE: reposiciona la secuencia del id, que los inserts con id explicito no\n"
        "avanzan. Sin esto la siguiente alta desde la app falla con clave duplicada:\n"
        "  select setval(\n"
        "    pg_get_serial_sequence('public.personal_bajas', 'id'),\n"
        "    (select max(id) from public.personal_bajas),\n"
        "    true\n"
        "  );"
    )


if __name__ == "__main__":
    main()

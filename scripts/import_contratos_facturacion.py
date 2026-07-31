"""Importa las cinco tablas maestras de facturacion desde Excel a Supabase.

Conserva los IDs de origen y usa upsert, por lo que se puede ejecutar de nuevo
sin duplicar filas.

Uso:
  python scripts/import_contratos_facturacion.py --dry-run
  python scripts/import_contratos_facturacion.py

Requiere SUPABASE_SERVICE_ROLE_KEY. La URL se toma de --url o del proyecto EDP.
"""
from __future__ import annotations

import argparse
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SOURCE_DIR = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"
DEFAULT_BATCH_SIZE = 250

TABLES = [
    {
        "table": "contratos_estado_facturas",
        "file": "tbl_contratos_estado_facturas.xlsx",
        "columns": {
            "Id_contrato_estado_factura": "id",
            "registro": "registro",
            "estado": "estado",
            "descripcion": "descripcion",
        },
        "integer": {"id", "registro"},
    },
    {
        "table": "contratos_fechas",
        "file": "tbl_contratos_fechas.xlsx",
        "columns": {
            "Id": "id",
            "contrato_id": "contrato_id",
            "fecha_inicio": "fecha_inicio",
            "fecha_fin": "fecha_fin",
            "concepto": "concepto",
        },
        "integer": {"id", "contrato_id"},
    },
    {
        "table": "contratos_presupuestos",
        "file": "tbl_contratos_presupuestos.xlsx",
        "columns": {
            "Id": "id",
            "contrato_id": "contrato_id",
            "periodo": "periodo",
            "Presupuesto": "presupuesto",
            "fecha_inicio": "fecha_inicio",
            "fecha_fin": "fecha_fin",
            "Observacion": "observacion",
            "Porcentaje_IVA": "porcentaje_iva",
        },
        "integer": {"id", "contrato_id"},
        "numeric": {"presupuesto", "porcentaje_iva"},
    },
    {
        "table": "contratos_funciones",
        "file": "tbl_contratos_funciones.xlsx",
        "columns": {
            "id_contrato_puesto": "id",
            "contrato_id": "contrato_id",
            "funcion_id": "funcion_id",
            "modalidad_id": "modalidad_id",
            "observacion": "observacion",
            "grupo": "grupo",
            "precio_01": "precio_01",
            "precio_02": "precio_02",
            "tipo_precio": "tipo_precio",
            "activo": "activo",
        },
        "integer": {"id", "contrato_id", "funcion_id", "modalidad_id"},
        "numeric": {"precio_01", "precio_02"},
        "boolean": {"activo"},
    },
    {
        "table": "contratos_facturacion",
        "file": "tbl_contratos_facturacion.xlsx",
        "columns": {
            "Id": "id",
            "fecha": "fecha",
            "T": "tipo",
            "serie": "serie",
            "n_documento": "n_documento",
            "Referencia": "referencia",
            "base_imponible": "base_imponible",
            "iva": "iva",
            "total": "total",
            "moneda": "moneda",
            "cod_cliente": "cod_cliente",
            "cliente": "cliente",
            "contrato_estado_factura_id": "contrato_estado_factura_id",
            "contrato_id": "contrato_id",
            "tipo_iva": "tipo_iva",
            "cobrada": "cobrada",
            "fecha_cobro": "fecha_cobro",
            "observacion": "observacion",
            "Presupuesto": "presupuesto_id",
            "unidades": "unidades",
            "precio_unitario": "precio_unitario",
        },
        "integer": {
            "id",
            "tipo",
            "contrato_estado_factura_id",
            "contrato_id",
            "presupuesto_id",
        },
        "numeric": {
            "base_imponible",
            "iva",
            "total",
            "tipo_iva",
            "unidades",
            "precio_unitario",
        },
        "boolean": {"cobrada"},
        "text": {"serie", "n_documento", "cod_cliente"},
    },
]


def normalize_value(value: Any, column: str, config: dict[str, Any]) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if column in config.get("boolean", set()):
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "sí", "si", "yes"}
    if column in config.get("integer", set()):
        return int(value)
    if column in config.get("numeric", set()):
        return float(Decimal(str(value)))
    if column in config.get("text", set()):
        return str(value)
    return value


def read_table(source_dir: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    path = source_dir / config["file"]
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    unknown = [header for header in headers if header not in config["columns"]]
    if unknown:
        raise ValueError(f"{path.name}: columnas no reconocidas: {unknown}")

    result = []
    for source_row in rows:
        if not any(value is not None for value in source_row):
            continue
        target_row = {}
        for header, value in zip(headers, source_row):
            target = config["columns"].get(header)
            if target:
                target_row[target] = normalize_value(value, target, config)
        result.append(target_row)
    return result


def api_request(
    url: str,
    key: str,
    path: str,
    *,
    method: str = "GET",
    body: Any = None,
    prefer: str = "",
) -> tuple[bytes, dict[str, str]]:
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(f"{url.rstrip('/')}/rest/v1/{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return response.read(), dict(response.headers)
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase devolvió HTTP {error.code}: {detail}") from error


def remote_ids(url: str, key: str, table: str) -> set[int]:
    raw, _ = api_request(url, key, f"{table}?select=id")
    return {int(row["id"]) for row in json.loads(raw or b"[]")}


def sanitize_catalog_references(
    data: dict[str, list[dict[str, Any]]],
    function_ids: set[int],
    modality_ids: set[int],
) -> None:
    for row in data["contratos_funciones"]:
        if row.get("funcion_id") is not None and row["funcion_id"] not in function_ids:
            print(
                f"Aviso: contratos_funciones {row['id']} referencia la función "
                f"inexistente {row['funcion_id']}; se importará sin función."
            )
            row["funcion_id"] = None
        if row.get("modalidad_id") is not None and row["modalidad_id"] not in modality_ids:
            print(
                f"Aviso: contratos_funciones {row['id']} referencia la modalidad "
                f"inexistente {row['modalidad_id']}; se importará sin modalidad."
            )
            row["modalidad_id"] = None


def validate(data: dict[str, list[dict[str, Any]]], contract_ids: set[int]) -> None:
    errors: list[str] = []
    budget_ids = {row["id"] for row in data["contratos_presupuestos"]}
    status_ids = {row["id"] for row in data["contratos_estado_facturas"]}

    for table, rows in data.items():
        missing_contracts = sorted({
            row["contrato_id"]
            for row in rows
            if row.get("contrato_id") is not None and row["contrato_id"] not in contract_ids
        })
        if missing_contracts:
            errors.append(f"{table}: contratos inexistentes {missing_contracts}")

    invoice_budget_ids = {
        row["presupuesto_id"]
        for row in data["contratos_facturacion"]
        if row.get("presupuesto_id") is not None
    }
    missing_budgets = sorted(invoice_budget_ids - budget_ids)
    if missing_budgets:
        errors.append(f"contratos_facturacion: presupuestos inexistentes {missing_budgets}")

    invoice_status_ids = {
        row["contrato_estado_factura_id"]
        for row in data["contratos_facturacion"]
        if row.get("contrato_estado_factura_id") is not None
    }
    missing_statuses = sorted(invoice_status_ids - status_ids)
    if missing_statuses:
        errors.append(f"contratos_facturacion: estados inexistentes {missing_statuses}")

    mismatched_budgets = []
    budgets_by_id = {row["id"]: row for row in data["contratos_presupuestos"]}
    for invoice in data["contratos_facturacion"]:
        budget = budgets_by_id.get(invoice.get("presupuesto_id"))
        if budget and invoice.get("contrato_id") != budget.get("contrato_id"):
            mismatched_budgets.append(invoice["id"])
    if mismatched_budgets:
        errors.append(f"facturas cuyo presupuesto pertenece a otro contrato: {mismatched_budgets}")

    if errors:
        raise ValueError("\n".join(errors))


def upsert_batches(
    url: str,
    key: str,
    table: str,
    rows: list[dict[str, Any]],
    batch_size: int,
) -> None:
    quoted_table = urllib.parse.quote(table)
    for start in range(0, len(rows), batch_size):
        batch = rows[start : start + batch_size]
        api_request(
            url,
            key,
            f"{quoted_table}?on_conflict=id",
            method="POST",
            body=batch,
            prefer="resolution=merge-duplicates,return=minimal",
        )
        print(f"  {table}: {min(start + len(batch), len(rows))}/{len(rows)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    parser.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.service_role_key:
        raise SystemExit("Falta SUPABASE_SERVICE_ROLE_KEY o --service-role-key.")

    data = {config["table"]: read_table(args.source_dir, config) for config in TABLES}
    contracts = remote_ids(args.url, args.service_role_key, "contratos")
    sanitize_catalog_references(
        data,
        remote_ids(args.url, args.service_role_key, "funciones"),
        remote_ids(args.url, args.service_role_key, "modalidades"),
    )
    validate(data, contracts)

    for table, rows in data.items():
        print(f"{table}: {len(rows)} filas validadas")
    if args.dry_run:
        print("Validación completada; no se ha escrito en Supabase.")
        return 0

    for config in TABLES:
        table = config["table"]
        upsert_batches(args.url, args.service_role_key, table, data[table], args.batch_size)
    print("Importación completada.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

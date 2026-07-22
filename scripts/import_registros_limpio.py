"""Recarga limpia de public.registros desde el Excel tbl_horas (formato 0722).

DIFERENCIAS CON import_registros_historico.py (que se conserva por trazabilidad
de la migracion original):

  - Las HORAS SE TOMAN TAL CUAL vienen en la columna `horas`. No se coalescen
    con hc/hf/hm ni se reconstruyen desde hora_fin - hora_inicio, aunque el
    intervalo no cuadre. En el origen nuevo la columna `horas` ya contiene las
    normales, complementarias, de montaje y de festivo trabajado; lo que
    distingue unas de otras es `tipo_hora_id`.
  - `horas_nocturnas` tambien se toma tal cual.
  - hc, hf, hm, hd, clases y horas_2 NO se importan (deben venir vacias; si
    traen algo se ignora y se reporta).
  - `bolsa_horas` SI se importa tal cual, para tratarla despues.
  - No se descarta ninguna fila por horas negativas ni por FTRAB ambiguo: se dan
    por buenas. Las anomalias solo se reportan.
  - `horas` nula se importa como 0 (la columna es real y se suma en el motor de
    nominas).

Se mantiene de la version anterior:
  - situacion_id 9 (codigo fuera del catalogo) -> NULL.
  - actividad_id NO se migra (los ids historicos no corresponden con
    public.actividades, que es la planificacion actual) -> NULL.
  - id_hora se guarda en legacy_id_hora y el insert es upsert por esa columna,
    asi que reejecutar no duplica.

Uso:
  python scripts/import_registros_limpio.py --dry-run
  python scripts/import_registros_limpio.py
  (requiere SUPABASE_SERVICE_ROLE_KEY en el entorno, o --service-role-key)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime, time as dtime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SOURCE_PATH = Path(r"C:\Users\Jr\OneDrive\JR\Tablas maestras\tbl_horas0722.xlsx")
DEFAULT_SUPABASE_URL = "https://epbtoarkinvgcaewbtvs.supabase.co"
DEFAULT_BATCH_SIZE = 2000

SITUACION_UNKNOWN = {9}

# Columnas del Excel que NO se importan porque deben venir vacias.
IGNORADAS = ["hc", "hf", "hm", "hd", "clases", "horas_2"]

# Columnas FK cuyo valor hay que validar contra los catalogos antes de cargar.
FK_COLUMNS = {
    "empresa_id": "empresa_id", "contrato_id": "contrato_id", "personal_id": "personal_id",
    "instalacion_id": "instalacion_id", "categoria_id": "categoria_id", "puesto_id": "puesto_id",
    "funcion_id": "funcion_id", "modalidad_id": "modalidad_id", "Nivel_id": "nivel_id",
    "grupo_id": "grupo_id", "dia_id": "dia_id", "tipo_hora_id": "tipo_hora_id",
    "situacion_id": "situacion_id",
}


def to_iso_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def to_iso_time(value):
    if value is None:
        return None
    if isinstance(value, dtime):
        return value.strftime("%H:%M:%S")
    return str(value)


def to_iso_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def to_num(value):
    return None if value is None else float(value)


def to_int(value):
    return None if value is None else int(value)


def fetch_ids(url: str, key: str, table: str) -> set[int]:
    """Ids de un catalogo, paginando (PostgREST corta en 1000)."""
    ids: set[int] = set()
    offset = 0
    while True:
        req = urllib.request.Request(
            f"{url.rstrip('/')}/rest/v1/{table}?select=id&order=id.asc&offset={offset}&limit=1000",
            method="GET",
        )
        req.add_header("apikey", key)
        req.add_header("Authorization", f"Bearer {key}")
        with urllib.request.urlopen(req, timeout=120) as resp:
            page = json.loads(resp.read().decode("utf-8"))
        if not page:
            break
        ids.update(int(x["id"]) for x in page)
        offset += len(page)
        if len(page) < 1000:
            break
    return ids


def load_and_transform(path: Path, limit: int | None = None, contratos_validos: set[int] | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else h for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    out: list[dict] = []
    avisos = {
        "horas_nulas": 0, "horas_negativas": 0, "situacion_desconocida": 0,
        "con_bolsa": 0, "ignoradas_con_valor": {c: 0 for c in IGNORADAS if c in idx},
        # contrato_id que no existe en el catalogo -> se importa NULL para no
        # romper la FK a mitad de carga. Se reporta para crearlo y reasignar.
        "contratos_desconocidos": {},
    }
    fk_valores: dict[str, set] = {dest: set() for dest in FK_COLUMNS.values()}

    for n, row in enumerate(rows):
        if row[idx["id_hora"]] is None:
            continue
        if limit is not None and n >= limit:
            break

        for c in avisos["ignoradas_con_valor"]:
            if row[idx[c]] not in (None, 0, ""):
                avisos["ignoradas_con_valor"][c] += 1

        horas = to_num(row[idx["horas"]])
        if horas is None:
            avisos["horas_nulas"] += 1
            horas = 0.0
        elif horas < 0:
            avisos["horas_negativas"] += 1

        situacion_id = to_int(row[idx["situacion_id"]])
        if situacion_id in SITUACION_UNKNOWN:
            avisos["situacion_desconocida"] += 1
            situacion_id = None

        bolsa = to_num(row[idx["bolsa_horas"]]) if "bolsa_horas" in idx else None
        if bolsa not in (None, 0):
            avisos["con_bolsa"] += 1

        contrato_id = to_int(row[idx["contrato_id"]])
        if (contrato_id is not None and contratos_validos is not None
                and contrato_id not in contratos_validos):
            avisos["contratos_desconocidos"][contrato_id] = (
                avisos["contratos_desconocidos"].get(contrato_id, 0) + 1)
            contrato_id = None

        registro = {
            "legacy_id_hora": to_int(row[idx["id_hora"]]),
            "fecha": to_iso_date(row[idx["fecha"]]),
            "actividad_id": None,
            "empresa_id": to_int(row[idx["empresa_id"]]) or 1,
            "contrato_id": contrato_id,
            "personal_id": to_int(row[idx["personal_id"]]),
            "instalacion_id": to_int(row[idx["instalacion_id"]]),
            "categoria_id": to_int(row[idx["categoria_id"]]),
            "puesto_id": to_int(row[idx["puesto_id"]]),
            "funcion_id": to_int(row[idx["funcion_id"]]),
            "modalidad_id": to_int(row[idx["modalidad_id"]]),
            "nivel_id": to_int(row[idx["Nivel_id"]]),
            "grupo_id": to_int(row[idx["grupo_id"]]),
            "nota": row[idx["Nota"]],
            "dia_id": to_int(row[idx["dia_id"]]),
            "hora_inicio": to_iso_time(row[idx["hora_inicio"]]),
            "hora_fin": to_iso_time(row[idx["hora_fin"]]),
            "horas": horas,
            "bolsa_horas": bolsa,
            "horas_diurnas": to_num(row[idx["horas_diurnas"]]),
            "horas_nocturnas": to_num(row[idx["horas_nocturnas"]]),
            "descanso": bool(row[idx["descanso"]]) if row[idx["descanso"]] is not None else False,
            "activo": bool(row[idx["activo"]]) if row[idx["activo"]] is not None else True,
            "festivo": bool(row[idx["festivo"]]) if row[idx["festivo"]] is not None else False,
            "sustitucion": bool(row[idx["sustitucion"]]) if row[idx["sustitucion"]] is not None else False,
            "facturar": bool(row[idx["facturar"]]) if row[idx["facturar"]] is not None else True,
            "abonar": bool(row[idx["abonar"]]) if row[idx["abonar"]] is not None else True,
            "tipo_hora_id": to_int(row[idx["tipo_hora_id"]]),
            "situacion_id": situacion_id,
            "anio": to_int(row[idx["Año"]]),
            "observacion": row[idx["observacion"]],
            "control": to_iso_datetime(row[idx["control"]]),
            "factura": row[idx["factura"]],
        }
        for src, dest in FK_COLUMNS.items():
            v = registro.get(dest)
            if v is not None:
                fk_valores[dest].add(v)
        out.append(registro)

    wb.close()
    return out, avisos, fk_valores


def post_batch(url: str, key: str, batch: list[dict]) -> None:
    payload = json.dumps(batch, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/registros?on_conflict=legacy_id_hora",
        data=payload, method="POST",
        headers={
            "apikey": key, "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            resp.read()
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"Supabase devolvio {exc.code}: {exc.read().decode('utf-8')[:600]}") from exc


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE_PATH)
    ap.add_argument("--url", default=os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL))
    ap.add_argument("--service-role-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    ap.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    contratos_validos = None
    if args.service_role_key:
        contratos_validos = fetch_ids(args.url, args.service_role_key, "contratos")
        print(f"Catalogo de contratos: {len(contratos_validos)} ids.")

    print(f"Leyendo {args.source} ...")
    filas, avisos, fk_valores = load_and_transform(
        args.source, limit=args.limit, contratos_validos=contratos_validos)

    print(f"Filas a importar: {len(filas)}")
    print(f"  horas nulas -> 0:        {avisos['horas_nulas']}")
    print(f"  horas negativas (se dan por buenas): {avisos['horas_negativas']}")
    print(f"  situacion 9 -> NULL:     {avisos['situacion_desconocida']}")
    print(f"  filas con bolsa_horas:   {avisos['con_bolsa']}")
    print("  columnas ignoradas que traian valor:")
    for c, n in avisos["ignoradas_con_valor"].items():
        print(f"     {c}: {n}")
    if avisos["contratos_desconocidos"]:
        print("  AVISO - contrato_id inexistente en el catalogo -> importado como NULL:")
        for cid, n in sorted(avisos["contratos_desconocidos"].items()):
            print(f"     contrato {cid}: {n} filas")
    print("  valores FK distintos por columna:")
    for dest, vals in sorted(fk_valores.items()):
        muestra = sorted(v for v in vals if v is not None)
        print(f"     {dest}: {len(muestra)} distintos, min={muestra[0] if muestra else '-'}, max={muestra[-1] if muestra else '-'}")

    if args.dry_run:
        print("\nDry-run: no se ha escrito nada.")
        return 0

    if not args.service_role_key:
        print("Falta SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 2

    total = len(filas)
    for start in range(0, total, args.batch_size):
        post_batch(args.url, args.service_role_key, filas[start:start + args.batch_size])
        print(f"  {min(start + args.batch_size, total)}/{total}", flush=True)

    print("Importacion terminada.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
